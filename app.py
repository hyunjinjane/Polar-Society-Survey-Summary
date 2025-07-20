import streamlit as st
import pandas as pd
import io
import datetime
import msoffcrypto
from openpyxl import load_workbook

# 숨겨진 행을 제거하며 읽기
def read_excel_skipping_hidden_rows(file_stream) -> pd.DataFrame:
    wb = load_workbook(file_stream, data_only=True)
    ws = wb.active

    # 첫 번째 행을 헤더로 사용
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]

    # 숨겨진 행 제외
    data = []
    for i, row in enumerate(ws.iter_rows(min_row=2), start=2):
        if not ws.row_dimensions[i].hidden:
            data.append([cell.value for cell in row])

    return pd.DataFrame(data, columns=headers)

st.set_page_config(page_title="건강검진 자가설문지 정리", layout="wide")
st.markdown("<h1 style='text-align: center;'>🏥 건강검진 자가설문지 정리</h1>", unsafe_allow_html=True)
st.write("엑셀 파일을 업로드하면, 줄바꿈 포함 정리 엑셀 파일을 다운받을 수 있습니다.")

uploaded_file = st.file_uploader("📤 엑셀 파일 업로드 (xlsx 형식)", type=["xlsx"])
password = st.text_input("🔐 파일에 비밀번호가 있을 경우 입력하세요 (없으면 비워두세요)", type="password")

if uploaded_file is not None:
    try:
        # 파일 읽기 및 암호 해제 처리
        if password:
            office_file = msoffcrypto.OfficeFile(uploaded_file)
            office_file.load_key(password=password)
            decrypted = io.BytesIO()
            office_file.decrypt(decrypted)
            decrypted.seek(0)
            df = read_excel_skipping_hidden_rows(decrypted)
        else:
            file_copy = io.BytesIO(uploaded_file.read())
            df = read_excel_skipping_hidden_rows(file_copy)

        # 정보 컬럼 탐색
        possible_info_cols = {
            '이름': 'Name',
            '생년월일': 'Date of Birth',
            '성별': 'SEX',
            '소속기관': 'Organization'
        }
        info_cols = []
        for kor, eng in possible_info_cols.items():
            if kor in df.columns:
                info_cols.append(kor)
            elif eng in df.columns:
                info_cols.append(eng)

        # 시작/끝 컬럼 자동 탐색
        start_idx = next((i for i, col in enumerate(df.columns) if '심장혈관계' in str(col) or 'Cardiovascular' in str(col)), None)
        end_idx = next((i for i, col in enumerate(df.columns) if str(col).startswith('40-1. 파견 장소') or str(col).startswith('40-1 Deployment Location')), None)

        if start_idx is None or end_idx is None:
            raise ValueError("설문 시작 또는 끝 컬럼을 찾을 수 없습니다.")

        survey_cols = df.columns[start_idx:end_idx]
        survey_df = df[survey_cols]

        results = []
        for i, row in survey_df.iterrows():
            info = df.loc[i, info_cols].to_dict()
            answered = row.notna().sum()
            not_answered = row.isna().sum()

            yes_values = ['예', 'YES', 'Yes', 'yes']
            no_values = ['아니오', 'NO', 'No', 'no']

            yes_cols = row[row.isin(yes_values)].index.tolist()
            etc_cols = row[~row.isin(yes_values + no_values) & row.notna()]
            etc_info = list(zip(etc_cols.index, etc_cols.values))

            combined = {
                **info,
                '총_답변수': answered,
                '미응답수': not_answered,
                "'아니오'_응답수": row.isin(no_values).sum(),
                "'예'_응답수": row.isin(yes_values).sum(),
                "'예'_응답_항목": '\n'.join(yes_cols),
                "'기타'_응답수": len(etc_info),
                "기타_응답": '\n'.join([f"{col} → {val}" for col, val in etc_info])
            }
            results.append(combined)

        summary_df = pd.DataFrame(results)
        summary_df.insert(0, '번호', range(1, len(summary_df) + 1))

        ordered_cols = ['번호'] + info_cols + [
            '총_답변수', '미응답수', "'아니오'_응답수",
            "'예'_응답수", "'예'_응답_항목",
            "'기타'_응답수", "기타_응답"
        ]
        summary_df = summary_df[[col for col in ordered_cols if col in summary_df.columns]]

        st.success("✅ 분석이 완료되었습니다! 아래에서 요약 데이터를 확인하고 파일을 다운로드하세요.")
        st.markdown(f"<h5>👥 총 설문자 수: <span style='color:#0066cc'>{len(summary_df)}명</span></h5>", unsafe_allow_html=True)
        st.dataframe(summary_df, use_container_width=True)

        st.markdown("## 📋 응답 요약 보기")
        for idx, row in summary_df.iterrows():
            name = row.get('이름') or row.get('Name', '이름 없음')
            birth = row.get('생년월일') or row.get('Date of Birth', '생년월일 없음')
            org = row.get('소속기관') or row.get('Organization', '소속 없음')

            st.markdown(f"""
                <h4 style='margin-bottom:0.2em;'>🔹 {idx+1}. <span style="color:#333;">{name} ({birth})</span></h4>
                <p style='margin-top:0; margin-bottom:0.5em;'>소속기관: <b>{org}</b></p>
            """, unsafe_allow_html=True)

            if row["'예'_응답_항목"]:
                st.markdown(f"""
                    <div style='background-color:#e6f4ea; padding:10px; border-radius:8px; margin-bottom:8px;'>
                    ✅ <b>'예' 응답 항목:</b><br>{row["'예'_응답_항목"].replace(chr(10), '<br>')}
                    </div>
                """, unsafe_allow_html=True)

            if row["기타_응답"]:
                st.markdown(f"""
                    <div style='background-color:#fdf3e6; padding:10px; border-radius:8px; margin-bottom:8px;'>
                    📝 <b>기타 응답:</b><br>{row['기타_응답'].replace(chr(10), '<br>')}
                    </div>
                """, unsafe_allow_html=True)

            st.markdown("<hr style='border:1px solid #ddd;'>", unsafe_allow_html=True)

        # 엑셀 저장 및 다운로드
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
            summary_df.to_excel(writer, index=False, sheet_name='응답요약')
            workbook = writer.book
            worksheet = writer.sheets['응답요약']
            wrap_format = workbook.add_format({'text_wrap': True})
            for col in ["'예'_응답_항목", "기타_응답"]:
                if col in summary_df.columns:
                    col_idx = summary_df.columns.get_loc(col)
                    worksheet.set_column(col_idx, col_idx, 60, wrap_format)

        st.download_button(
            label="📥 엑셀 파일 다운로드",
            data=output.getvalue(),
            file_name=f"설문_요약_{datetime.date.today()}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except Exception as e:
        st.error(f"⚠️ 오류가 발생했습니다: {e}")
